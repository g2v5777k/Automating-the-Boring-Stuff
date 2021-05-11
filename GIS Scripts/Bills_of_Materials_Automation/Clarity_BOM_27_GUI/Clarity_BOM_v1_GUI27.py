import os
from os import path
import getpass
from openpyxl import load_workbook
import PySimpleGUI27 as sg
from site import addsitedir
from sys import executable
import time, os, sys, math, traceback



# trash to attempt to get gui to compile properly
##################################################################
# List of clarity features
clarity_features = ['Anchors', 'FiberCable', 'conduit', 'OLT_Boundaries', 'Structures', 'SpliceClosure', 'Slackloops', 'Strand']

# Workbook vars
cwd = os.getcwd()
# template = os.path.join(cwd + '\\template', 'Clarity_BOM_Template.xlsx')
template = os.path.join(cwd, 'Clarity_BOM_Template.xlsx')
wb = load_workbook(filename=template)
ws = wb['Sheet1']

# sg initialization
interpreter = executable
path_to_interpreter = path.dirname(interpreter)
sitepkg = path_to_interpreter + "\\site-packages"
addsitedir(sitepkg)
import arcpy # Apparently need to put the import arcpy statement after the sg initialization for desktop app to work.

gui = [[sg.Text('Input GDB', size=(20, 1)), sg.InputText(size=(70, 1)), sg.FolderBrowse()],
       [sg.Text('BOM Output Location:', size=(20, 1)), sg.InputText(size=(70, 1)), sg.FolderBrowse()],
       [sg.Text('OLT(s) Comma Separated:', size=(20, 1)), sg.InputText(size=(70, 1))],
       [sg.Submit(), sg.Exit()]]

window = sg.Window('Clarity BOM (Simplified)').Layout(gui)
scratchExists = os.path.exists(os.path.join(r'C:\Users', getpass.getuser(), 'Documents', 'ArcGIS','scratch.gdb'))
if scratchExists == False:
    arcpy.CreateFileGDB_management(os.path.join(r'C:\Users', getpass.getuser(), 'Documents', 'ArcGIS'), 'scratch.gdb')
else:
    pass
scratch = os.path.join(r'C:\Users', getpass.getuser(), 'Documents', 'ArcGIS','scratch.gdb')

##################################################################

def clear_scratch():
    # add something to skip deleting these and transferring fcs in the fcs are already present (speeds up process)

    arcpy.env.workspace = scratch
    fc_list = arcpy.ListFeatureClasses()
    fc_tables = arcpy.ListTables()
    for fc in fc_list:
        arcpy.Delete_management(fc)
    for table in fc_tables:
        arcpy.Delete_management(table)

def transfer_fcs():
    # transfer clarity fcs to a scratch gdb
    arcpy.env.workspace = input_gdb
    for fc in clarity_features:
        arcpy.FeatureClassToGeodatabase_conversion(fc, scratch)

def get_olt_features(olt):
    '''
    Transfer necessary fcs to a scratch.gdb workspace, then ingest an OLT string, pull all features contained within that OLT from parent Clarity GDB.
    :return:
    Create fiber (aerial and ug), strand/conduit, scs, peds, vaults, slackloops, and anchors fcs for use in below functions
    '''

    # get olt, pull just boundary
    arcpy.env.workspace = scratch
    olt_exp = "name = '{}'".format(olt)
    arcpy.Select_analysis('OLT_Boundaries', 'olt', olt_exp)

    # pull features within olt boundary
    arcpy.Intersect_analysis(['FiberCable', 'olt'], 'olt_fiber', '', '', 'LINE')
    arcpy.Intersect_analysis(['conduit', 'olt'], 'olt_conduit', '', '', 'LINE')
    arcpy.Intersect_analysis(['Strand', 'olt'], 'olt_strand', '', '', 'LINE')
    arcpy.SpatialJoin_analysis('Anchors', 'olt', 'olt_anchors', 'JOIN_ONE_TO_ONE', 'KEEP_COMMON', '', 'COMPLETELY_WITHIN')
    arcpy.SpatialJoin_analysis('Structures', 'olt', 'olt_structures', 'JOIN_ONE_TO_ONE', 'KEEP_COMMON', '', 'COMPLETELY_WITHIN')
    arcpy.SpatialJoin_analysis('SpliceClosure', 'olt', 'olt_scs', 'JOIN_ONE_TO_ONE', 'KEEP_COMMON', '', 'COMPLETELY_WITHIN')
    arcpy.SpatialJoin_analysis('Slackloops', 'olt', 'olt_slackloops', 'JOIN_ONE_TO_ONE', 'KEEP_COMMON', '', 'COMPLETELY_WITHIN')

def fiber():
    '''
    Calculate footage for fiber in requirement number #2. Calculations are identical.
    :return:
    Write footage to template.
    '''

    # UG
    exp = "subtypecode = 1 AND fiber_built = 'N'"
    ugFiber = scratch + '\\' + 'ug_fiber'
    arcpy.Select_analysis(scratch + '\\' + 'olt_fiber', ugFiber, exp)
    values = [row[0] for row in arcpy.da.SearchCursor(ugFiber, ['subtypecode'])]
    if not values:
        ws['B2'] = 0
    else:
        length = 0
        arcpy.Dissolve_management(ugFiber, scratch + '\\' + 'ugFiber_Dissolve', 'subtypecode', )
        arcpy.AddGeometryAttributes_management(scratch + '\\' + 'ugFiber_Dissolve', 'LENGTH_GEODESIC', 'FEET_US')
        sum_length = [length + row[0] for row in arcpy.da.SearchCursor(scratch + '\\' + 'ugFiber_Dissolve', ['LENGTH_GEO'])]
        ws['B2'] = sum_length[0]

    # Aerial
    exp = "subtypecode = 2 AND fiber_built = 'N'"
    aerialFiber = scratch + '\\' + 'aerial_fiber'
    arcpy.Select_analysis(scratch + '\\' + 'olt_fiber', aerialFiber, exp)
    values = [row[0] for row in arcpy.da.SearchCursor(aerialFiber, ['subtypecode'])]
    if not values:
        ws['B3'] = 0
    else:
        length = 0
        arcpy.Dissolve_management(aerialFiber, scratch + '\\' + 'Aerial_Dissolve', 'subtypecode', )
        arcpy.AddGeometryAttributes_management(scratch + '\\' + 'Aerial_Dissolve', 'LENGTH_GEODESIC', 'FEET_US')
        sum_length = [length + row[0] for row in arcpy.da.SearchCursor(scratch + '\\' + 'Aerial_Dissolve', ['LENGTH_GEO'])]
        ws['B3'] = sum_length[0]

def strand_conduit():
    '''
    Calculate footage for strand and conduit in requirement #2. Calculations are identical.
    :return:
    Write footage to template
    '''

    # Strand, calculated from fiber field where strand_built = 'N'
    exp = "subtypecode = 2 AND strand_built = 'N'"
    strandFiber = scratch + '\\' + 'strandFiber'
    arcpy.Select_analysis(scratch + '\\' + 'olt_fiber', strandFiber, exp)
    values = [row[0] for row in arcpy.da.SearchCursor(strandFiber, ['subtypecode'])]
    if not values:
        lenStrand = 0
    else:
        length = 0
        arcpy.Dissolve_management(strandFiber, scratch + '\\' + 'strandFiber_Dissolve', 'subtypecode', )
        arcpy.AddGeometryAttributes_management(scratch + '\\' + 'strandFiber_Dissolve', 'LENGTH_GEODESIC', 'FEET_US')
        sum_length = [length + row[0] for row in arcpy.da.SearchCursor(scratch + '\\' + 'strandFiber_Dissolve', ['LENGTH_GEO'])]
        lenStrand = sum_length[0]

    # Conduit, calculate from ugFiber where strand_built = 'N'
    exp = "subtypecode = 1 AND strand_built = 'N'"
    strandConduit = scratch + '\\' + 'strandConduit'
    arcpy.Select_analysis(scratch + '\\' + 'olt_fiber', strandConduit, exp)
    values = [row[0] for row in arcpy.da.SearchCursor(strandConduit, ['subtypecode'])]
    if not values:
        lenConduit = 0
    else:
        length = 0
        arcpy.Dissolve_management(strandConduit, scratch + '\\' + 'strandConduit_Dissolve', 'subtypecode', )
        arcpy.AddGeometryAttributes_management(scratch + '\\' + 'strandConduit_Dissolve', 'LENGTH_GEODESIC', 'FEET_US')
        sum_length = [length + row[0] for row in arcpy.da.SearchCursor(scratch + '\\' + 'strandConduit_Dissolve', ['LENGTH_GEO'])]
        lenConduit = sum_length[0]

    ws['B4'] = lenStrand + lenConduit

def naps():
    '''
    Count splice closures in requirement #3
    :return:
    Write count to template
    '''

    exp = "spliceuse = 'NAP'"
    naps = scratch + '\\' + 'naps'
    arcpy.Select_analysis(scratch + '\\' + 'olt_scs', naps, exp)
    count = arcpy.GetCount_management(naps)
    ws['B5'] = count[0]

def peds_vaults():
    '''
    Count pedestals and vaults in requirement #4
    :return:
    Write count to template
    '''

    # Pedestals, subtypecode = 5 then count features
    exp = "subtypecode = 5"
    peds = scratch + '\\' + 'peds'
    arcpy.Select_analysis(scratch + '\\' + 'olt_structures', peds, exp)
    count = arcpy.GetCount_management(peds)
    ws['B7'] = count[0]

    # Vaults subtypecode = 4 then count features
    exp = "subtypecode = 4"
    vaults = scratch + '\\' + 'vaults'
    arcpy.Select_analysis(scratch + '\\' + 'olt_structures', vaults, exp)
    count = arcpy.GetCount_management(vaults)
    ws['B6'] = count[0]

def slack_anchors():
    '''
    Count slackloops and anchors within boundary in requirement #5
    :return:
    Write counts to template
    '''

    slackCount = arcpy.GetCount_management(scratch + '\\' + 'olt_slackloops')
    anchorCount = arcpy.GetCount_management(scratch + '\\' + 'olt_anchors')
    ws['B8'] = slackCount[0]
    ws['B9'] = anchorCount[0]

def main(input_olts):
    olts = input_olts.replace(' ', '').split(',')
    for olt in olts:
        print('Prepping Workspace...\n')
        clear_scratch()
        transfer_fcs()
        print('\n')
        print('Creating BOM for OLT: ' + olt + '\n')
        get_olt_features(olt)
        fiber()
        strand_conduit()
        naps()
        peds_vaults()
        slack_anchors()
        fn = output + "\\{0}_BOM_{1}.xlsx".format(olt, str(time.strftime("%Y%m%d")))
        wb.save(fn)
        print('BOM Created.')


if __name__ == '__main__':
    # Created simple gui application for user ease of use. Initialized in config.py.
    while True:
        try:
            event, values = window.Read()
            if event is None or event == 'Exit':
                break
            input_gdb = values[0]
            output = values[1]
            olts = values[2]
            main(olts)
            sg.Popup('BOM(s) created.')
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            sg.Popup('Error in creating BOM(s)!', "Error : " + str(e), "LINE NO: {0}".format(exc_tb.tb_lineno))


