import arcpy
import pickle
import os
import math
from os.path import join as join
from googleapiclient.http import MediaIoBaseDownload
import getpass
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import io
import zipfile
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from sys import executable
import traceback
from logging_decorator import makelogger,logError

root = os.path.dirname(os.path.abspath(__file__))
inputs = join(root, 'input')
outpath = join(root, 'output')
logDir = join(root, 'logs')
if os.path.exists(join(root, 'scratch.gdb')):
    pass
else:
    arcpy.CreateFileGDB_management(root, 'scratch.gdb')
scratch = join(root, 'scratch.gdb')
SCOPES = ['https://www.googleapis.com/auth/drive']
date = datetime.now()
gdb = 'RDOF_Design_{}.zip'.format(date.strftime('%Y%m%d'))
gdbpath = os.path.join(inputs, gdb).strip('.zip') + '.gdb'
datasetPath = gdbpath + '\\' + 'RDOF_Design'
template = join(root, 'RDOF_BOM_Template.xlsx')
wb = load_workbook(filename=template)
ws = wb['Aspire']

necessaryFCs = ['Proposed_OLT_LCP_Boundaries',
                'ServedAddress',
                'DropFiber',
                'Conduit',
                'FiberCable',
                'SpliceClosure',
                'FiberEquipment',
                'Structure',
                'SlackLoop',
                'Riser',
                'Proposed_Cabinets',
                'RDOF_CBG',
                'OVERBUILD_POLY']