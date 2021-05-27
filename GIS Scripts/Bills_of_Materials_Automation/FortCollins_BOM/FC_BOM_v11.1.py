'''
created by - Matt Huhmann
created on - 2/12/2019
Version History
v1 - started
v2 - added GUI to create button for design team to run - unknown
v3 - trouble shooting gui and gdb download issues - cluttrell
v4 - exe troubleshooting - cluttrell
v5 - Fixed issues with huts > 99 not producing data - cluttrell
v8 - Adding in comments fields for conduit calcs - cluttrell
v9 - Added in new params for cells A21, A25-A29 subsequent row vars were updated - cluttrell
v10 - Fixed an issue where 288ct fiber was not populating in the BOM templates. Fixed an issue where 'pvault' values were being incorrectly attributed.
      Change was to explicitly select only preliminary values to avoid misattribution - cluttrell
v11 - Fixed an issue with a SplicePoint query that did not select features whose name was equal to that of the fdh. This query was deleted- cluttrell
v11.1 - Added a line (164) to encode (utf-8) FDH strings as users were getting encoding errors. - cluttrell
'''
import time, os, sys, math, traceback, getpass
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
import PySimpleGUI27 as sg

# The following code helps in creating a stand-alone executable
from site import addsitedir
from sys import executable
from os import path

interpreter = executable
path_to_interpreter = path.dirname(interpreter)
sitepkg = path_to_interpreter + "\\site-packages"
# print "The path to site-package is..", sitepkg
addsitedir(sitepkg)

import arcpy

fiber_fields = ["fdhid", "cablename", "fibercount", "fdhcable", "length_geo"]
con_fields = ["fdhid", "layer", "diameter", "mid_id", "length_geo", "shared", 'installmethod']
sc_fields = ["fdhid", "locationdescription", "splice_type", "fcount", "splice_count", "fiber_assignments", "sc_size"]
vault_fields = ["fdhid", "layer", "structure_type", "pvault", 'interconnect']
trench_fields = ["fdhid", "length_geo"]
loop_fields = ["fdhid", "designid", "p_hierarchy", "cablecapacity", "measuredlength"]

layout = [[sg.Text('Fill in the fields below', size=(45, 1))],
          [sg.Text('Input GDB:', size=(20, 1)), sg.InputText(size=(70, 1)), sg.FolderBrowse()],
          [sg.Text('Output Location:', size=(20, 1)), sg.InputText(size=(70, 1)), sg.FolderBrowse()],
          [sg.Text('FDH list(comma separted):', size=(20, 1)), sg.InputText(size=(70, 1))],
          [sg.Submit(), sg.Exit()]]

window = sg.Window('Fort Collins BOM Tool').Layout(layout)

codeblock = '''
def getfdh(x):
	if x is not None and "-" in x:
                test = len(x.split("-")[1])
                holder = x.split("-")
                if test > 4:
                        fdh = holder[0] + "-" + holder[1][:4]

                if test <= 4:
                        fdh = holder[0] + "-" + holder[1][:3]
                return fdh
        else:
                return ""
'''


def get_end_points(x, y, location):
    """
    This function is used to get the end points of the fibercable and then
    create a new feature class with all those points.
    PARAMETERS
    ----------
    x : feature class(complete path)
        feature class for which you need end points
    y : feature class(complete path)
        Name of the new feature class that has to be created
    location : string
        There are two options
        'BOTH_ENDS' : If you need both the end points of a line, you can use this option
        'MID' : If you need just the mid-point of a line, you can use this option
    """

    fiber = []
    fieldnames = [field.name for field in arcpy.ListFields(x) if field.name not in ["shape_Length"]]
    fieldnames.append("SHAPE@")
    with arcpy.da.SearchCursor(x, fieldnames) as cur:
        for row in cur:
            temp = list(row)
            fiber.append(temp)

    cur_gdb, fc = os.path.split(y)
    arcpy.CreateFeatureclass_management(cur_gdb, fc, "POINT", x, "", "", x)

    insert_cursor = arcpy.da.InsertCursor(y, fieldnames)
    index = 0
    arrayObj = arcpy.Array()
    for row in fiber:
        if location == "BOTH_ENDS":
            pntObj = arcpy.Point()
            pntObj.ID = index
            pntObj.X = row[-1].firstPoint.X
            pntObj.Y = row[-1].firstPoint.Y
            index += 1
            temp = row[:]
            temp[-1] = pntObj
            insert_cursor.insertRow(temp)

            pntObj = arcpy.Point()
            pntObj.ID = index
            pntObj.X = row[-1].lastPoint.X
            pntObj.Y = row[-1].lastPoint.Y
            index += 1
            temp = row[:]
            temp[-1] = pntObj
            insert_cursor.insertRow(temp)
        elif location == "MID":
            pntObj = arcpy.Point()
            pntObj.ID = index
            pntObj.X = row[-1].positionAlongLine(0.50, True).firstPoint.X
            pntObj.Y = row[-1].positionAlongLine(0.50, True).firstPoint.Y
            index += 1
            temp = row[:]
            temp[-1] = pntObj
            insert_cursor.insertRow(temp)

    del insert_cursor


def prep_data(fdhs):
    """
    This function helps in data preparation
    PARAMETERS
    ----------
    fdhs : list
        A list of FDH IDs for which you need the BOMs
    """
    crs = arcpy.SpatialReference(2231)
    arcpy.env.overwriteOutput = True
    arcpy.env.workspace = scratch
    print("FILTERING DATA")

    """
    If there are any Feature classes or Tables present in the scratch GDB,
    remove all of them
    """
    fcs = arcpy.ListFeatureClasses()
    for fc in fcs:
        arcpy.Delete_management(scratch + '/' + fc)
    tables = arcpy.ListTables()
    for table in tables:
        arcpy.Delete_management(scratch + '/' + table)

    # The keys present in the following dictionary are the feature classes
    # Data from these feature classes are gathered to generate BOM
    # and the values are the attributes present in those feature classes.
    # These attributes are later(lines 147 - 166) used in filtering the data
    name_dict = {'FiberLine': 'cablename', 'FC_Structure': 'layer',
                 'StructureLine': 'layer', 'fdhpoint': 'fdhid', 'SplicePoint': 'locationdescription',
                 'FiberSlackLoop': 'designid'
                 }

    # The following fdh expression helps in generating a query of below form
    # ("fdhid" = 'DIX101d-F31' or "fdhid" = 'DIX101d-F32' or "fdhid" = 'DIX101d-F33')
    # which can later be used to select only the required FD Boundaries
    fdh_exp = "(" + " or ".join(["fdhid = '{0}'".format(x) for x in fdhs]) + ")"
    fdh_exp.encode('utf-8').strip()
    # Select only those FDH Boundaries for which the BOMs needs to be generated
    arcpy.Select_analysis(gdb + "\\fdhboundary", scratch + "\\fdhs", fdh_exp)

    """ Exlanations for Queries used inside select_analysis for the for loop part that comes next

    # Query for Structure and Conduit
    # 		Select only those structures and conduits for which the status is 'Preliminary'
    #       and the ones which are present inside the FDH Boundaries we are working on (This part is
    #       handled using Intersect_analysis)
    # Then for the next elif part, the queries are much similar to the above queries and so are self explanatory
    # Same goes for final else part
    """

    for fc in name_dict.keys():  # ["FiberOpticCable", "FC_Structure", "FC_Conduit", "fdhpoint", "SplicePoint"]:
        fieldnames = [field.name for field in arcpy.ListFields(gdb + "\\" + fc)]

        if fc == "SplicePoint":
            arcpy.Select_analysis(gdb + "\\" + fc, scratch + "\\" + fc)

        elif fc in ['FC_Structure', 'StructureLine']:
            arcpy.Select_analysis(gdb + "\\" + fc, scratch + "\\temp_" + fc, "inventory_status_code = 'Preliminary'")
            arcpy.Intersect_analysis([scratch + "\\temp_" + fc, scratch + "\\fdhs"], scratch + "\\" + fc)
        elif "inventory_status_code" in fieldnames:
            arcpy.Select_analysis(gdb + "\\" + fc, scratch + "\\pre_" + fc,
                                  "(" + " or ".join(["{0} like '{1}%'".format(name_dict[fc], x) for x in
                                                     fdhs]) + ") and inventory_status_code = 'Preliminary'")
            arcpy.Select_analysis(gdb + "\\" + fc, scratch + "\\" + fc,
                                  "(" + " or ".join(["{0} like '{1}%'".format(name_dict[fc], x) for x in fdhs]) + ")")
            arcpy.AddField_management(scratch + "\\pre_" + fc, "fdhid", "TEXT")
            arcpy.CalculateField_management(scratch + "\\pre_" + fc, "fdhid", "getfdh(!{0}!)".format(name_dict[fc]),
                                            "PYTHON_9.3", codeblock)
        else:
            arcpy.Select_analysis(gdb + "\\" + fc, scratch + "\\" + fc,
                                  "(" + " or ".join(["{0} like '{1}%'".format(name_dict[fc], x) for x in fdhs]) + ")")

        # Make sure there is an 'fdhid' column for all of the feature classes.
        # There is no special reason for this. It's just to make some of the other geo-processing operations faster
        fieldnames = [field.name for field in arcpy.ListFields(scratch + "\\" + fc)]
        if "fdhid" not in fieldnames:
            arcpy.AddField_management(scratch + "\\" + fc, "fdhid", "TEXT")
            arcpy.CalculateField_management(scratch + "\\" + fc, "fdhid", "getfdh(!{0}!)".format(name_dict[fc]),
                                            "PYTHON_9.3", codeblock)

    # Select only Access Fiber, changed 12/07 to grab all fiber intersecting an FDH, and included 'Lateral' infrastructure class query so that 288 cts are counted.
    arcpy.Intersect_analysis([gdb + "\\FiberLine", scratch + "\\fdhs"], scratch + "\\af_1", '', '', 'LINE')
    arcpy.Select_analysis(scratch + "\\af_1", scratch + "\\af", "infrastructureclass = 'Access' OR infrastructureclass = 'Lateral'")

    # Get the end points of the Access Fiber
    get_end_points(scratch + "\\af", scratch + "\\af_ends", "BOTH_ENDS")

    # Get those fiber ends which intersects with Splice Point
    arcpy.SpatialJoin_analysis(scratch + "\\SplicePoint", scratch + "\\af_ends", scratch + "\\af_sc_join",
                               "JOIN_ONE_TO_MANY", "KEEP_ALL", "", "INTERSECT", "")

    # We dissolve the output from previous step just to make sure we have only one entry even for the points where multiple fibercable intersect with a splice point
    # We will take into consideration only the fibercable with maximum fiber count. Thats the reason why we use ["fibercount", "MAX"]

    arcpy.Dissolve_management(scratch + "\\af_sc_join", scratch + "\\final_scs",
                              ["locationdescription", "splice_type", "splice_count", "fdhid", "fiber_assignments",
                               "spliceenclosuremodelnumber"], [["fibercount", "MAX"]])  # "cable_size",
    arcpy.AlterField_management(scratch + "\\final_scs", "MAX_fibercount", "fcount", "fcount")
    arcpy.AlterField_management(scratch + "\\final_scs", "spliceenclosuremodelnumber", "sc_size", "sc_size")

    # The below set of lines (220- 227) are used to create a feature class with name final_vaults
    # A new attribute named 'pvault' is added and it's value is either 'Y' or 'N' - Changed 12/07/2020 to only include preliminary structures pvault = 'N'

    # Added prelim_vaults 12/07/2020
    arcpy.Select_analysis(gdb + "\\FC_Structure", scratch + "\\prelim_vaults", "inventory_status_code = 'Preliminary'")

    arcpy.AddField_management(scratch + "\\FC_Structure", "pvault", "TEXT")
    arcpy.MakeFeatureLayer_management(scratch + "\\FC_Structure", "vaults")
    # arcpy.CalculateField_management("vaults", "pvault", "'N'", "PYTHON_9.3", "")
    arcpy.SelectLayerByLocation_management("vaults", "INTERSECT", scratch + "\\prelim_vaults", "", "NEW_SELECTION")
    arcpy.CalculateField_management("vaults", "pvault", "'N'", "PYTHON_9.3", "")
    arcpy.SelectLayerByAttribute_management("vaults", "CLEAR_SELECTION")
    arcpy.CopyFeatures_management("vaults", scratch + "\\final_vaults")

    # The following set of lines(234 - 240) are used to find out whether an access fiber cable is an FDH cable.
    # Any Acces Fibercable that intersects FDH point is an 'FDH cable.'
    # So, we add a new field named 'fdhcable' and it's values are 'Y' or 'N'
    # If the value is 'Y' - it means fiber is an FDH Cable else it is not.
    # And the final result is copied into scratch GDB just like vaults

    arcpy.AddField_management(scratch + "\\af", "fdhcable", "TEXT")
    arcpy.MakeFeatureLayer_management(scratch + "\\af", "fiber")
    arcpy.SelectLayerByLocation_management("fiber", "INTERSECT", scratch + "\\fdhpoint", "", "NEW_SELECTION")
    arcpy.CalculateField_management("fiber", "fdhcable", "'Y'", "PYTHON_9.3", "")
    arcpy.SelectLayerByAttribute_management("fiber", "CLEAR_SELECTION")
    arcpy.CopyFeatures_management("fiber", scratch + "\\final_fiber")

    arcpy.AddGeometryAttributes_management(scratch + "\\final_fiber", "LENGTH_GEODESIC", "FEET_US", "", crs)

    arcpy.Select_analysis(scratch + "\\StructureLine", scratch + "\\all_con",
                          "diameter = '2inch' or diameter = '1.25inch'")
    arcpy.AddField_management(scratch + "\\all_con", "shared", "TEXT")
    arcpy.CalculateField_management(scratch + "\\all_con", "shared", "'N'", "PYTHON_9.3", "")
    arcpy.SplitLine_management(scratch + "\\all_con", scratch + "\\con_split")
    get_end_points(scratch + "\\con_split", scratch + "\\con_mids", "MID")
    arcpy.AddField_management(scratch + "\\con_mids", "trench", "SHORT")
    arcpy.CalculateField_management(scratch + "\\con_mids", "trench", "1", "PYTHON_9.3", "")
    arcpy.Buffer_analysis(scratch + "\\con_mids", scratch + "\\con_mid_buff", "1.5 FEET", "FULL", "ROUND")
    arcpy.Dissolve_management(scratch + "\\con_mid_buff", scratch + "\\con_mid_diss", "", "", "SINGLE_PART", "")
    arcpy.AddField_management(scratch + "\\con_mid_diss", "mid_id", "LONG")
    arcpy.CalculateField_management(scratch + "\\con_mid_diss", "mid_id", "!objectid!", "PYTHON_9.3", "")
    arcpy.SpatialJoin_analysis(scratch + "\\con_mid_buff", scratch + "\\con_mid_diss", scratch + "\\con_join_temp",
                               "JOIN_ONE_TO_ONE", "KEEP_ALL", "", "INTERSECT", "")
    arcpy.Dissolve_management(scratch + "\\con_join_temp", scratch + "\\con_mid_diss_temp", ["mid_id"],
                              [["trench", "SUM"]], "SINGLE_PART", "")
    arcpy.AlterField_management(scratch + "\\con_mid_diss_temp", "SUM_trench", "trench", "trench")
    arcpy.SpatialJoin_analysis(scratch + "\\con_split", scratch + "\\con_mid_diss_temp", scratch + "\\con_join",
                               "JOIN_ONE_TO_ONE", "KEEP_ALL", "", "INTERSECT", "")

    arcpy.Select_analysis(scratch + "\\con_join", scratch + "\\con2", "diameter = '2inch'")
    arcpy.Select_analysis(scratch + "\\con_join", scratch + "\\con125", "diameter = '1.25inch'")
    arcpy.Buffer_analysis(scratch + "\\con2", scratch + "\\con2_buff", "2 FEET", "FULL", "ROUND", "ALL")
    arcpy.MakeFeatureLayer_management(scratch + "\\con125", "con125")
    arcpy.SelectLayerByLocation_management("con125", "WITHIN", scratch + "\\con2_buff", "", "NEW_SELECTION")
    arcpy.CalculateField_management("con125", "shared", "'Y'", "PYTHON_9.3", "")
    arcpy.SelectLayerByAttribute_management("con125", "CLEAR_SELECTION")
    arcpy.Merge_management([scratch + "\\con2", "con125"], scratch + "\\final_con")
    arcpy.AddGeometryAttributes_management(scratch + "\\final_con", "LENGTH_GEODESIC", "FEET_US", "", crs)

    arcpy.Dissolve_management(scratch + "\\final_con", scratch + "\\trench", ["fdhid"])
    arcpy.AddGeometryAttributes_management(scratch + "\\trench", "LENGTH_GEODESIC", "FEET_US", "", crs)

    print("DATA FILTERATION DONE..")


def read_data(fdh):
    """
    Read all the data from all the final feature classes that are created from the
    data preparation part.
    This function is called for each individual FDH Boundary at a time as opposed
    to the data preparation part which is done for all the FDH Boundaries at a time
    (Explanation for the lines 235 - 243)
    Using Search Cursor, we iterate through all the features of a feature class
    corresponding to a particular FDH.
    Any of the field value which is None replaced by empty space('') and the domain coded
    values are replaced by proper domain descriptions and if there are any text values
    that contain complete numerical values, those are converted to int.
    Each output returned is a list of lists and each list contains
    the data of a row(in a feature class)
     """
    print("READING FILTERED DATA")
    arcpy.env.workspace = scratch
    fiber = []
    scs = []
    conduit = []
    vaults = []
    trench = []

    fcs = ["final_fiber", "final_scs", "final_con", "final_vaults", "trench"]
    cur_fields = [fiber_fields, sc_fields, con_fields, vault_fields, trench_fields]
    out_data = [fiber, scs, conduit, vaults, trench]
    # out_data = []
    i = 0
    domains = arcpy.da.ListDomains(scratch)
    for fc, req_fields, out_fc in zip(fcs, cur_fields, out_data):
        # print "Read FINAL FC: {0}".format(fc)
        out_data.append([])
        allfields = arcpy.ListFields(fc)
        field_map = {o.name.lower(): o for o in allfields}
        fields = [field_map[x.lower()] for x in req_fields]
        with arcpy.da.SearchCursor(scratch + "\\" + fc, [x.name for x in fields],
                                   "fdhid like '{0}'".format(fdh)) as cursor:
            for row in cursor:
                print(row)
                temp = ["" if x is None else (x if fields[ind].domain == "" else
                                              [y for y in domains if y.name == fields[ind].domain][0].codedValues[x])
                        for ind, x in enumerate(row)]
                for i, z in enumerate(temp):
                    try:
                        temp[i] = float(z)
                    except:
                        temp[i] = z
                out_fc.append(temp)
        #print("Read FINAL FC: {0} len: {1}".format(fc, len(out_fc)))

    return (fiber, scs, conduit, vaults, trench)


def uptwelve(x, base=5):
    """
    This function is used for some mathematical calculation which is used to fill cell 'A35'
    as per the BOM guidelines (which I'm not aware of at the moment)
    """
    return int(base * math.ceil(float(x) / base))


def createbom(fdh, fiber, scs, conduit, vaults, trench):
    wb = load_workbook(filename=template)
    ws = wb['AEG Units']
    print("Creating BOM for {0}".format(fdh))

    '''
    fiber_fields = ["fdhid", "cablename", "fibercount", "fdhcable", "length_geo"]
    con_fields = ["fdhid", "layer", "diameter", "mid_id", "length_geo"]
    sc_fields = ["fdhid", "locationdescription", "splice_type", "fcount", "splice_count", "fiber_assignments", "sc_size"]
    vault_fields = ["fdhid", "layer", "structure_type", "pvault"]
    trench_fields = ["fdhid", "length_geo"]
    loop_fields = ["fdhid", "designid", "p_hierarchy", "cablecapacity", "measuredlength"]
    '''

    # Missile Bore - All 1.25" except where shared path with 2"
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if
                              x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "1.25inch" and
                              x[con_fields.index("installmethod")] == "Missile" and x[con_fields.index("installmethod")] != "Directional-Parallel"])))
    # print('1.25 missle bore', temp)
    ws['A7'] = temp

    # Directional Bore - up to 2" Conduit = All 2" Conduit, plus shared path with 1.25"
    seen = set()
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if
                              x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "2inch" or x[con_fields.index("installmethod")] == 'Directional-Parallel'])))
    # print('directional, shared 1.25 and 2', temp)
    ws['A8'] = temp  # number of conduit in trench?

    # Low Density Downtown Bore
    temp = 0  # int(math.ceil(sum([x[trench_fields.index("length_geo")] for x in trench if x[trench_fields.index("fdhid")] == fdh])))
    ws['A10'] = temp

    # High Density Downtown Open Cut
    temp = 0
    ws['A11'] = temp

    # UG "Special Crossing" (RR, Interstate, Waterway) - individual xing pricing will vary
    temp = 0
    ws['A12'] = temp

    # Conduit Adder in same trench (1st additional conduit - up to 2" Conduit) Parallel
    seen = set()
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if
                              x[con_fields.index("fdhid")] == fdh and x[con_fields.index("installmethod")] == 'Directional-Parallel'])))
    # print('Conduit adder in same trench, up to 2 inch parallel', temp)
    ws['A13'] = temp

    # Underground Rear Easement Adder - New Conduit
    temp = 0
    ws['A19'] = temp

    # Install Vault (17x30x24)
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Small Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == ''])
    ws['A20'] = temp

    # Install Vault (24x36x24, intermediate)
    temp = len([x for x in vaults if x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Intermediate Vault" and x[vault_fields.index('interconnect')] == ''])
    ws['A21'] = temp

    # Install Vault (30x48x36)
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Medium Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == ''])
    ws['A22'] = temp

    # Install Vault (36x60x36)
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Large Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == ''])
    ws['A23'] = temp

    # Install Flower Pot (10x10)
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Flower Pot" and x[vault_fields.index('interconnect')] == ''])
    ws['A24'] = temp

    # Install Vault (17x30x24) - adjacent to power vault
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Small Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == 'YES'])
    ws['A25'] = temp

    # Install Vault ((24x36x24, intermediate) - adjacent to power vault
    temp = len([x for x in vaults if x[vault_fields.index('fdhid')] == fdh and x[vault_fields.index("structure_type")] == "Intermediate Vault" and
                                       x[vault_fields.index('interconnect')] == 'YES'])
    ws['A26'] = temp

    # Install Vault (30x48x36) - adjacent to power vault
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Medium Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == 'YES'])
    ws['A27'] = temp

    # Install Vault (36x60x36) - adjacent to power vault
    temp = len([x for x in vaults if
                x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Large Vault" and
                x[vault_fields.index("pvault")] == "N" and x[vault_fields.index('interconnect')] == 'YES'])
    ws['A28'] = temp

    # Install Flower Pot (10x10) - adjacent to power vault
    temp = 0  # len([x for x in vaults if x[vault_fields.index("fdhid")] == fdh and x[vault_fields.index("structure_type")] == "Flower Pot" and x[vault_fields.index("pvault")] == "Y" x[vault_fields.index('interconnect')] == 'YES'])
    ws['A29'] = temp

    # Underground Rear Easement Adder - New Vault
    temp = 0
    ws['A30'] = temp

    # OTDR Testing and Documentation - Uni-directional LCP to NAP
    # print([x[sc_fields.index("fiber_assignments")] for x in scs if x[sc_fields.index("fdhid")] == fdh])
    try:
        temp = max([(int(x[sc_fields.index("fiber_assignments")].split("-")[-1]) if x[sc_fields.index(
            "fiber_assignments")] != 0 and
                                                                                    x[sc_fields.index(
                                                                                        "fiber_assignments")] and
                                                                                    x[sc_fields.index(
                                                                                        "fiber_assignments")] != '' else 0)
                    for x in scs if x[sc_fields.index("fdhid")] == fdh])
        temp = uptwelve(temp, 12) + 24
    except:
        temp = "None"
    ws['A37'] = temp

    # HANDLE THE CABLE LENGTH AND MULTIPLIER TABLE
    range = ws['G40:L45']
    csizes = [288, 144, 96, 48, 24, 12]
    for row, csize in zip(range, csizes):
        # print(csize)
        seen = set()
        row[0].value = len([x for x in fiber if
                            x[fiber_fields.index("fdhid")] == fdh and x[fiber_fields.index("fdhcable")] == 'Y' and x[
                                fiber_fields.index("fibercount")] == csize and x[
                                fiber_fields.index("cablename")] not in seen and
                            not seen.add(x[fiber_fields.index("cablename")])])

        row[1].value = len([x for x in scs if
                            x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("fcount")] == csize and "MCA" in x[
                                sc_fields.index("splice_type")]])
        row[2].value = len([x for x in scs if
                            x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("fcount")] == csize and "NAP" in x[
                                sc_fields.index("splice_type")] and "MCA" not in x[sc_fields.index("splice_type")]])
        row[3].value = len([x for x in scs if
                            x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("fcount")] == csize and "RE" in x[
                                sc_fields.index("splice_type")]])
        row[5].value = int(math.ceil(sum([x[fiber_fields.index("length_geo")] for x in fiber if
                                          x[fiber_fields.index("fdhid")] == fdh and x[
                                              fiber_fields.index("fibercount")] == csize])))
        # print(row[5].value)
    '''
    #Parallel 1.25"
    seen = set()
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "1.25inch" and x[con_fields.index("shared")] <> 'Y' and
                              ((x[con_fields.index("mid_id")] in seen and not seen.add(x[con_fields.index("mid_id")])) or seen.add(x[con_fields.index("mid_id")]))])))
    ws['G46'] = temp
    #SHARED 1.25" CONDUIT
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "1.25inch" and x[con_fields.index("shared")] == 'Y'])))
    ws['G47'] = temp
    #Parallel 2"
    seen = set()
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "2inch" and ((x[con_fields.index("mid_id")] in seen and not seen.add(x[con_fields.index("mid_id")])) or seen.add(x[con_fields.index("mid_id")]))])))
    ws['G48'] = temp
    '''
    # 1.25 inch CONDUIT
    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if
                              x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "1.25inch"])))
    # print('1.25 inch conduit', temp)
    ws['G54'] = temp

    # 2 inch CONDUIT

    temp = int(math.ceil(sum([x[con_fields.index("length_geo")] for x in conduit if
                              x[con_fields.index("fdhid")] == fdh and x[con_fields.index("diameter")] == "2inch"])))
    # print('2 inch conduit', temp)
    ws['G55'] = temp

    # Heat Shrink Sleeves = 1 per splice
    temp = int(math.ceil(sum(
        [(int(x[sc_fields.index("splice_count")]) if x[sc_fields.index("splice_count")] != '' else 0) for x in scs if
         x[sc_fields.index("fdhid")] == fdh])))
    ws['A66'] = temp

    # FOSC 450 A-Gel Splice Enclosure (used on 48F, 24F
    temp = len([x for x in scs if x[sc_fields.index("fdhid")] == fdh and x[
        sc_fields.index("sc_size")] == "Commscope FOSC 450 A-Gel" and "X" not in x[
                    sc_fields.index("locationdescription")]])
    ws['A67'] = temp

    # FOSC 450 A-Gel (used at NAP Extender solution only)
    temp = len([x for x in scs if x[sc_fields.index("fdhid")] == fdh and x[
        sc_fields.index("sc_size")] == "Commscope FOSC 450 A-Gel" and "X" in x[sc_fields.index("locationdescription")]])
    ws['A68'] = temp

    # FOSC 450 B-Gel Splice Enclosure
    temp = len([x for x in scs if
                x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("sc_size")] == "Commscope FOSC 450 B-Gel"])
    ws['A70'] = temp

    # FOSC 450 C-Gel Splice Enclosure
    temp = len([x for x in scs if
                x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("sc_size")] == "Commscope FOSC 450 C-Gel"])
    ws['A72'] = temp

    # FOSC 450 D-Gel Splice Enclosure
    temp = len([x for x in scs if
                x[sc_fields.index("fdhid")] == fdh and x[sc_fields.index("sc_size")] == "Commscope FOSC 450 D-Gel"])
    ws['A74'] = temp

    fn = Output_Location + "\\{0}_BOM_{1}.xlsx".format(fdh, str(time.strftime("%Y%m%d")))
    if os.path.exists(fn):
        os.remove(fn)
    wb.save(fn)


def main():
    # fdh_string can be a single fdh_id or multiple fdh_ids that are comma-separated values
    fdhs = fdh_string.replace(' ', '').split(',')

    # All the data corresponding to these FDHs is pulled and processes
    prep_data(fdhs)

    # Data for each individual FDH is gathered and BOM is created(one at a time)
    for fdh in fdhs:
        (fiber, scs, conduit, vaults, trench) = read_data(fdh)
        createbom(fdh, fiber, scs, conduit, vaults, trench)

    print("BOMS CREATED. HIT EXIT TO QUIT OR PUT IN NEW INFORMATION TO RUN DIFFERENT SHEETS")


while True:
    try:
        event, values = window.Read()
        if event is None or event == 'Exit':
            break

        gdb = values[0]
        scratch = os.path.join(r'C:\Users', getpass.getuser(), 'Documents', 'ArcGIS','scratch.gdb')  # abspath(r'C:\Users\ColtonLuttrell\Desktop\Desktop_Files\Python\Market_Specific_Scripts\FortCollins\scratch.gdb') #(r'C:\Users', getpass.getuser(), 'Documents', 'ArcGIS', 'scratch.gdb')
        cur_dir = os.path.join(os.path.dirname('__file__'))
        # cur_dir = r'C:\Users\ColtonLuttrell\Desktop\Desktop_Files\Python\Market_Specific_Scripts\FortCollins'
        # Place this template in the same location as your python file.
        # Also, place this template inside dist folder if you are creating a executable
        template = os.path.join(cur_dir, 'DIX101a-F02_BOM_20190529.xlsx')

        # This is where the BOM gets stored
        Output_Location = values[1]

        fdh_string = values[2]  # Example 'DIX101d-F31' or 'DIX101d-F31, DIX101d-F32, DIX101d-F33'
        main()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()  # This helps in finding the information about the exception.
        print("Error : " + str(e))  # This helps in finding out what the error is
        print("LINE NO: {0}".format(
            exc_tb.tb_lineno))  # This helps in finding out which line in the try block raised the error
        traceback.print_tb(exc_tb)  # This helps in tracing back the exception
        print("FAILURE! ONE OR MORE BOMS MAY NOT HAVE BEEN CREATED")

window.Close()
